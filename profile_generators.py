#%%
import os
from datetime import datetime, time, timedelta
from typing import Tuple, List, Union
from random import uniform, randint, seed, randrange, random
from numpy import nan, cos, pi, linspace, around, ravel, isnan
from numpy.random import normal, permutation

from numpy.random import seed as numpy_seed
from numpy.random import beta as beta_distribution
from pandas import *
from pandas.core.common import flatten
from pandas import IndexSlice as IDX
from utils import get_profile_data_from_excel

import pickle
import openpyxl
import math

import matplotlib.pyplot as plt
from matplotlib.pyplot import hist
import matplotlib.cbook

from plot import plot_FDP_input_preview

import enlopy

import warnings

warnings.filterwarnings("ignore", category=matplotlib.cbook.mplDeprecation)

""" _____________________________________________________________  HELPER FUNCTIONS """


def completely_unconstrained_profile(
    start: datetime, end: datetime, resolution: timedelta
) -> DataFrame:

    """ Can be used as a base model for device profiles. """

    df = DataFrame(
        columns=[
            "equals",
            "max",
            "min",
            "derivative equals",
            "derivative max",
            "derivative min",
        ],
        index=date_range(start=start, end=end, freq=resolution),
    )

    df.index.name = "datetime"

    return df


def limited_capacity_profile(
    start: datetime, end: datetime, resolution: timedelta, capacity: float
) -> DataFrame:

    """ Can be used to model a prosumer or a battery with unlimited storage capacity. """

    df = completely_unconstrained_profile(start=start, end=end, resolution=resolution)
    df["derivative max"] = capacity
    df["derivative min"] = -capacity

    return df


def limited_production_profile(
    start: datetime, end: datetime, resolution: timedelta, capacity: float
) -> DataFrame:

    """ Can be used to model a generator. """

    df = limited_capacity_profile(
        start=start, end=end, resolution=resolution, capacity=capacity
    )
    df["derivative max"] = 0

    return df


def time_plus(start_time: time, duration: timedelta):

    """ Returns a start datetime plus a given timedelta (Used for buffer profile generator) """

    start = datetime(
        2000,
        1,
        1,
        hour=start_time.hour,
        minute=start_time.minute,
        second=start_time.second,
    )
    end = start + duration

    return end.time()


def time_duration(start_time: time, end_time: time):

    """ Returns time duration between given datetimes (Used for buffer profile generator) """

    start_datetime = datetime(
        2000,
        1,
        1,
        hour=start_time.hour,
        minute=start_time.minute,
        second=start_time.second,
    )
    end_datetime = datetime(
        2000, 1, 1, hour=end_time.hour, minute=end_time.minute, second=end_time.second
    )
    duration = end_datetime - start_datetime
    return duration


def add_noise(
    inputs_values, freq: timedelta, mode, st, r=0.9, Lmin=0, force_timed_index=False
):
    """ 
    Add noise with given characteristics.

    Parameters:
    -------
        
    inputs_values : Series/DataFrame 
        1d or 2d timeseries
    
    mode : int
        1 Normal Distribution, 2: Uniform Distribution, 3: Gauss Markov (autoregressive gaussian)
    
    st : float 
        Noise parameter. Scaling of random values
    
    r : float 
        Applies only for mode 3. Autoregressive coefficient AR(1). Has to be between  [-1,1]
    
    Lmin : float
        minimum load values. This is used to trunc values below zero if they are generated with a lot of noise

    freq: timedelta
        Used within python enlopy.clean_convert

    Returns: 
    -------
        inputs_values with noise : Series
    
    """

    L = np.atleast_2d(inputs_values)

    if st == 0:
        print("No noise to add")
        return inputs_values

    loadlength = L.shape  # 8760
    if mode == 1:  # Normal
        noisevector = st * np.random.randn(
            *loadlength
        )  
        out = L * (1 + noisevector)
    elif mode == 2:  # Uniform
        noisevector = st * np.random.rand(*loadlength)
        out = L * ((1 - st) + st * noisevector)
    elif mode == 3:  
        out = enlopy.gen_gauss_markov(L, st, r)
    else:
        raise ValueError("Not available mode")
    out[out < Lmin] = Lmin  

    out = enlopy.clean_convert(
        np.squeeze(out), force_timed_index=force_timed_index, freq=freq
    )

    return out


""" _____________________________________________________________  PROFILE GENERATORS """


def create_fdp_inputs(
    start: datetime = None,
    end: datetime = None,
    resolution: timedelta = None,
    seed_value: int = None,
    upward_capacity_max: float = None,
    downward_capacity_max: float = None,
    std: float = None,
    autocorrelation: float = None,
    market_prices_noise_parameter: Tuple = None,
    shape_weight: float = None,
    result_directory: str = None,
    excel_file_location: ExcelFile = None,
    excel_file_name: str = None,
    # pickle_profile: bool=False,
    show_preview: bool = False,
    pickled_profile: str = None,
    repeat_daily: bool = False,
) -> None:

    """ 
        Creates the inputs for the flexibility demanding party and 
        stores it as an excel sheets for the scenario input file.
        Imbalance profiles gets created randomly using gauss-markov value generator from python enlopy module. 
        Market prices gets derived from the imbalance profile, after being shuffled and noised.
        Deviation prices gets derived from the market prices, and should be always greater than the market prices.
        NOTE: FDP uses deviation prices to hedge against market prices, if the aggregator fails to deliver, at least the market prices gets repaid.

        Parameters
        ----------
        start, end, resolution : datetime
            Simulation time related parameter

        seed_value : int
            Used for seed generators

        upward_capacity_max, downward_capacity_max: float
            Controls the span of the imbalances values, but also the distribution of positive/negative values
            in array_of_means for the gauss-markov random value generator
            
        std, autocorrelation : float
            Used for gauss-markov random value generator
            autocorrelation value needs to between -1 and 1
        
        market_prices_noise_parameter : tuple
            Sets the mean and the std for noising the created imbalance values when creating market prices

        shape_weight : float
            Defines how much the market prices correlates with the magnitudes of the imbalance values
            Must be between 0 and 1

        input_file: str
            Contains path to excel input file

        fdp_profile_path: str
            Contains path to pickled profile

        Returns
        -------
        None
    """

    # Assertations
    assert (
        autocorrelation >= -1 and autocorrelation <= 1
    ), "Autocorrelation value must be between -1 and 1."
    assert (
        shape_weight >= 0 and shape_weight <= 1
    ), "Shape weight value must be between 0 and 1."

    # Set simulation index from attributes
    simulation_time = date_range(start=start, end=end, freq=resolution)

    if pickled_profile is None:

        # Flag
        repeat = repeat_daily

        # Create full simulation time dataframe
        fdp_full_time_profile = DataFrame(
            index=simulation_time,
            columns=[
                "imbalances",
                "market_prices",
                "deviation_price_up",
                "deviation_price_down",
            ],
        )

        fdp_full_time_profile.index.name = "datetime"
        excel_file = os.path.splitext(excel_file_location)[0].split("/")[-1] + ".xlsx"

        if os.path.isfile(excel_file_location):

            fdp_profile = get_profile_data_from_excel(
                entity="FDP",
                excel_file_location=excel_file_location,
                simulation_time=simulation_time,
            )

            # Infer and assign index resolution
            fdp_profile.index.freq = infer_freq(fdp_profile.index)

            assert (
                to_timedelta(fdp_profile.index.freq) == resolution
            ), "Resolution not equal"

            # Set repeat if length is not equal
            if len(fdp_profile) < len(simulation_time):
                repeat = True

            if repeat:

                delta_profile = fdp_profile.index[-1] - fdp_profile.index[0]

                if delta_profile > timedelta(days=1):
                    fdp_profile = fdp_profile.loc[start : start + timedelta(days=1)]

                simulation_days = (end - start).days

                if not delta_profile < timedelta(days=1):

                    delta_profile = timedelta(days=1)

                for i in range(0, simulation_days):
                    day_start = start + timedelta(days=i)
                    fdp_full_time_profile.loc[
                        day_start : day_start + delta_profile
                    ] = fdp_profile.loc[start : start + delta_profile].values

                fdp_profile = fdp_full_time_profile

        else:

            print("Creating excel input file")

            # Set full year index for python enlopy value generator
            full_year_index = date_range(
                start=datetime(year=2018, month=1, day=1),
                end=datetime(year=2019, month=1, day=1, hour=0),
                freq=resolution,
            )
            # Drop last index (00:00)
            full_year_index = full_year_index.drop(full_year_index[-1])

            # Fix random parameters
            numpy_seed(seed_value)
            seed(seed_value)

            # Define arguments for enlopy.gen_gauss_markov value function
            periods = len(full_year_index)
            array_of_means = [
                randrange(start=-downward_capacity_max, stop=upward_capacity_max)
                for p in range(periods)
            ]
            array_of_std = [std for p in range(periods)]

            # Create imbalances profile for one year
            imbalances = enlopy.gen_gauss_markov(
                array_of_means, array_of_std, autocorrelation
            )

            # Assure that there are more positive than negative imbalances
            if imbalances[imbalances < 0].sum() * -1 > imbalances[imbalances > 0].sum():
                imbalances *= -1

            # Convert imbalances to series
            imbalances = Series(index=full_year_index, data=imbalances)

            # Create market price data based on imbalance values
            market_prices = abs(imbalances.copy())

            # Shuffle to add randomness
            market_prices_shuffled = (
                market_prices.copy().sample(frac=1).reset_index(drop=True)
            )

            # Define noise parameters for additional randomness on market prices
            mu, sigma = market_prices_noise_parameter
            market_prices_noise = list(
                flatten(normal(mu, sigma, [len(market_prices_shuffled.index), 1]))
            )

            market_prices_shuffled = market_prices_shuffled + market_prices_noise
            market_prices_shuffled.index = full_year_index

            # Use weights to define how much the market prices correlate with the imbalance values
            weight_1 = shape_weight
            weight_2 = 1 - weight_1

            market_prices = market_prices.copy() * shape_weight + market_prices_shuffled.copy() * (
                1 - shape_weight
            )

            # Randomize the market prices to get deviation prices
            market_prices_noised = Series(
                index=full_year_index,
                data=[
                    round(price * uniform(0.1, 10)) for price in market_prices.copy()
                ],
            )

            deviation_prices = Series(
                index=full_year_index,
                # data=[price * uniform(1, 10) for price in market_prices_noised.values],
                data=100,
            )

            # Slice yearly profiles according to simulation time
            imbalances = Series(index=simulation_time, data=around(imbalances, 2))
            market_prices_noised = Series(
                index=simulation_time, data=around(market_prices_noised.loc[start:end])
            )
            deviation_prices = Series(
                index=simulation_time, data=around(deviation_prices.loc[start:end])
            )

            # Assign profiles to input dataframe for excel
            fdp_full_time_profile[
                "imbalances"
            ] = imbalances  
            fdp_full_time_profile["market_prices"] = (
                market_prices_noised / 10
            ) 
            fdp_full_time_profile[
                "deviation_price_up"
            ] = deviation_prices  
            fdp_full_time_profile["deviation_price_down"] = (
                deviation_prices * -1
            )  

            if repeat:

                delta_profile = (
                    fdp_full_time_profile.index[-1] - fdp_full_time_profile.index[0]
                )

                simulation_days = (end - start).days

                if not delta_profile < timedelta(days=1):

                    delta_profile = timedelta(days=1)

                for i in range(0, simulation_days):
                    day_start = start + timedelta(days=i)
                    fdp_full_time_profile.loc[
                        day_start : day_start + delta_profile
                    ] = fdp_full_time_profile.loc[start : start + delta_profile].values

                fdp_profile = fdp_full_time_profile

    # # Pickled profile passed
    else:
        fdp_profile = pickle.load(open(pickled_profile, "rb"))

        assert len(fdp_profile) >= len(
            fdp_profile
        ), "Pickled profile does not match simulation runtime"

    # Create a pickle profile
    pickle.dump(fdp_profile, open(result_directory + "/PROFILES/FDP_Inputs.p", "wb"))

    # Store profile back to the excel file
    if os.path.isfile(excel_file_location):

        wb = openpyxl.load_workbook(excel_file_location)
        std = wb.get_sheet_by_name("FDP")
        wb.remove_sheet(std)
        wb.save(filename=excel_file_location)

        with ExcelWriter(excel_file_location, engine="openpyxl", mode="a") as writer:
            fdp_profile.to_excel(writer, sheet_name="FDP")
    else:

        # Create a Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(filename=excel_file_location)

        with ExcelWriter(excel_file_location, engine="openpyxl", mode="a") as writer:
            fdp_profile.to_excel(writer, sheet_name="FDP")

    writer.save()
    writer.close()

    if show_preview:
        # Plot profiles
        fig = plt.figure(1, figsize=(20, 10))
        fig.suptitle("Flexibility demanding inputs", fontsize=14)

        ax1 = plt.subplot(3, 1, 1)
        fdp_profile["imbalances"].plot(
            kind="Bar", title="Imbalances", grid=True, ax=ax1
        )
        plt.setp(ax1.get_xticklabels(), visible=False)
        plt.xticks([])

        ax2 = plt.subplot(3, 1, 2)
        fdp_profile["market_prices"].plot(
            kind="Line", c="blue", ax=ax2, title="Market prices", grid=True, sharex=ax1
        )
        plt.setp(ax2.get_xticklabels(), visible=False)

        ax3 = plt.subplot(3, 1, 3)
        fdp_profile["deviation_price_up"].plot(
            kind="Line",
            c="orange",
            ax=ax3,
            title="Deviation prices",
            grid=True,
            sharex=ax1,
        )
        fdp_profile["deviation_price_down"].plot(
            kind="Line",
            c="red",
            ax=ax3,
            title="Deviation prices",
            grid=True,
            sharex=ax1,
        )

        plt.show()

    # Basic stats
    total_days = (end - start).days
    day_cnt = 0


    return


#%%
def create_beta_distributed_inputs(
    start,
    end,
    resolution,
    episodes,
    seed_value: int = None,
    result_directory: str = None,
    excel_file_location: ExcelFile = None,
    excel_file_name: str = None,
    imbalances: bool = False,
    market_prices: bool = False,
    scenario_name: str = None,
) -> None:

    episode_index = list(range(1, episodes + 1))
    # Set simulation index from attributes
    simulation_time = date_range(start=start, end=end - resolution, freq="D")
    simulation_daytime = date_range(start=start, end=end - resolution, freq=resolution)

    unique_dates = simulation_time.normalize().unique()

    one_day_index = date_range(
        start=datetime(year=2018, month=1, day=1),
        end=datetime(year=2018, month=1, day=2, hour=0) - resolution,
        freq=resolution,
    )

    if imbalances:

        # Fix random parameters
        numpy_seed(seed_value)
        seed(seed_value)

        profile = DataFrame(index=simulation_daytime, columns=["Values"], data=0)

        profile.index.name = "Datetime"
        profile["Episode"] = 0

        # print('fdp_orders: ', fdp_orders)
        for enum, day in enumerate(unique_dates):
            # Get episode number
            profile.loc[day : day + timedelta(days=1), "Episode"] = enum + 1

        profile.set_index("Episode", append=True, inplace=True)
        profile = profile.swaplevel(axis=0)

        one_day_index = date_range(
            start=datetime(year=2018, month=1, day=1),
            end=datetime(year=2018, month=1, day=2, hour=0) - resolution,
            freq=resolution,
        )

        imbalances_profile = DataFrame(
            index=simulation_time, columns=[list(range(0, len(one_day_index)))],
        )

        daily_profile = DataFrame(
            index=one_day_index,
            # columns=["imbalances", "market_prices"],
            # data=0
        )

        steps = linspace(-1, 1, len(daily_profile.index))

        # # FDP_INPUT 1
        # trig_function = [math.sin(x*1.7-5) + 2.3 + math.cos(x*12+1) for x in steps]
        # daily_profile["Values"] = trig_function
        # alpha = [np.random.randint(2,5) for x in one_day_index]
        # beta = [4 for x in one_day_index]
        # for enum,index in enumerate(simulation_time):
        #     dist = beta_distribution(a=alpha, b=beta, size=len(one_day_index))
        #     # noise = np.random.randint(np.random.randint(-2,0) , np.random.randint(0,5), size=(len(dist),))
        #     noise = 0
        #     values = around(daily_profile["Values"].values + dist +  noise,0)
        #     values[11:] = 0
        #     imbalances_profile.loc[index] = values

        # FDP_INPUT 2
        trig_function = [
            math.sin(x * 1.7 - 5) + 2.3 + math.cos(x * 23 + 8) for x in steps
        ]
        # trig_function = [3 for x in steps]
        daily_profile["Values"] = trig_function
        alpha = [np.random.uniform(2, 5) for x in one_day_index]
        beta = [4 for x in one_day_index]

        for enum, index in enumerate(simulation_time):
            dist = beta_distribution(a=alpha, b=beta, size=len(one_day_index))
            noise = np.random.uniform(-1, 1, size=(len(dist),))
            # noise = 0
            # values = around(daily_profile["Values"].values + noise, 1)
            values = around(daily_profile["Values"].values + dist + noise, 2)
            # values = around(dist +  noise, 1)
            values[11:] = 0
            imbalances_profile.loc[index] = values
            imbalances_profile.loc[index] = imbalances_profile.loc[index].clip(lower=0)

        # plt.plot(steps, trig_function, color = 'red', marker = "o")
        # plt.plot(steps, daily_profile["Values"].values, color = 'red', marker = "o")
        # plt.title("Imbalances trig function")
        # plt.xlabel("time")
        # plt.ylabel("Y")
        # plt.show()

        ib_stats = DataFrame(
            index=[
                "Mean",
                "Median",
                "Upper Quartile",
                "Lower Quartile",
                "IQR",
                "Upper Whisker",
                "Lower Whisker",
                "Max",
                "Min",
            ],
            columns=[list(range(0, len(one_day_index)))],
        )

        columns = [one_day_index.strftime("%H:%M")]

        for column in imbalances_profile.columns:

            data = imbalances_profile[column]

            ib_stats.loc["Mean", column] = round(np.mean(data), 2)
            ib_stats.loc["Median", column] = round(np.median(data), 2)
            ib_stats.loc["Upper Quartile", column] = round(np.percentile(data, 75), 2)
            ib_stats.loc["Lower Quartile", column] = round(np.percentile(data, 25), 2)
            ib_stats.loc["IQR", column] = round(
                np.percentile(data, 75) - np.percentile(data, 25), 2
            )
            ib_stats.loc["Upper Whisker", column] = round(
                data[
                    data <= np.percentile(data, 75) + 1.5 * ib_stats.loc["IQR", column]
                ].max(),
                2,
            )
            ib_stats.loc["Lower Whisker", column] = round(
                data[
                    data >= np.percentile(data, 25) - 1.5 * ib_stats.loc["IQR", column]
                ].min(),
                2,
            )
            ib_stats.loc["Max", column] = round(data.max(), 2)
            ib_stats.loc["Min", column] = round(data.min(), 2)

        with open(result_directory + "/PROFILES/FDP_imbalances.txt", "a") as f:
            f.write(ib_stats.to_string(header=True, index=True))

    if market_prices:

        # Fix random parameters
        numpy_seed(seed_value + 1)
        seed(seed_value + 1)

        profile = DataFrame(index=simulation_daytime, columns=["Values"], data=0)

        profile.index.name = "Datetime"
        profile["Episode"] = 0

        # print('fdp_orders: ', fdp_orders)
        for enum, day in enumerate(unique_dates):
            # Get episode number
            profile.loc[day : day + timedelta(days=1), "Episode"] = enum + 1

        profile.set_index("Episode", append=True, inplace=True)
        profile = profile.swaplevel(axis=0)

        market_price_profile = DataFrame(
            index=simulation_time, columns=[list(range(0, len(one_day_index)))],
        )

        daily_profile = DataFrame(
            index=one_day_index,
        )

        steps = linspace(-1, 1, len(daily_profile.index))

        trig_function = [math.sin(x * 1.7) + 2.3 + math.cos(x * 22 + 6) for x in steps]
        daily_profile["Values"] = trig_function

        for enum, index in enumerate(simulation_time):
            alpha = [np.random.uniform(1, 5) for x in one_day_index]
            beta = [np.random.uniform(1, 5) for x in one_day_index]
            dist = beta_distribution(a=alpha, b=beta, size=len(one_day_index))
            noise = np.random.uniform(-0.5, 0.5, size=(len(dist),)) * np.random.randint(
                -1, 1
            )
            values = around(daily_profile["Values"].values + dist + noise, 2)
            values[11:] = 0
            market_price_profile.loc[index] = values
            market_price_profile.loc[index] = market_price_profile.loc[index].clip(
                lower=0.1
            )

        mp_stats = DataFrame(
            index=[
                "Mean",
                "Median",
                "Upper Quartile",
                "Lower Quartile",
                "IQR",
                "Upper Whisker",
                "Lower Whisker",
                "Max",
                "Min",
            ],
            columns=[list(range(0, len(one_day_index)))],
        )

        columns = [one_day_index.strftime("%H:%M")]

        for column in market_price_profile.columns:

            data = market_price_profile[column]

            mp_stats.loc["Mean", column] = round(np.mean(data), 2)
            mp_stats.loc["Median", column] = round(np.median(data), 2)
            mp_stats.loc["Upper Quartile", column] = round(np.percentile(data, 75), 2)
            mp_stats.loc["Lower Quartile", column] = round(np.percentile(data, 25), 2)
            mp_stats.loc["IQR", column] = round(
                np.percentile(data, 75) - np.percentile(data, 25), 2
            )
            mp_stats.loc["Upper Whisker", column] = round(
                data[
                    data <= np.percentile(data, 75) + 1.5 * mp_stats.loc["IQR", column]
                ].max(),
                2,
            )
            mp_stats.loc["Lower Whisker", column] = round(
                data[
                    data >= np.percentile(data, 25) - 1.5 * mp_stats.loc["IQR", column]
                ].min(),
                2,
            )
            mp_stats.loc["Max", column] = round(data.max(), 2)
            mp_stats.loc["Min", column] = round(data.min(), 2)

        with open(
            "C:/Users/W/Desktop/ComOpt/comopt/simulator/scenarios/case_studies/master_thesis/FDP_PROFILES/FDP_.txt",
            "a",
        ) as f:
            f.write(mp_stats.to_string(header=True, index=True))

    plot_FDP_input_preview(
        start=start,
        end=end,
        episodes=episodes,
        imbalances_profile=imbalances_profile,
        imbalances_stats=ib_stats,
        market_price_profile=market_price_profile,
        market_price_stats=mp_stats,
        one_day_index=one_day_index,
        result_directory=result_directory,
        scenario_name=scenario_name,
    )

    # Write to pickle
    updated_profile = pickle.load(
        open(result_directory + "/PROFILES/FDP_Inputs.p", "rb")
    )

    if market_prices:
        market_price_profile = market_price_profile.T.squeeze()
        values = [
            market_price_profile[column] for column in market_price_profile.columns
        ]
        values = ravel(values)
        updated_profile["market_prices"].iloc[:-1] = values

    if market_prices:
        imbalances_profile = imbalances_profile.T.squeeze()
        values = [imbalances_profile[column] for column in imbalances_profile.columns]
        values = ravel(values)
        updated_profile["imbalances"].iloc[:-1] = values

    # Save to excel
    wb = openpyxl.load_workbook(excel_file_location)
    std = wb.get_sheet_by_name("FDP")
    wb.remove_sheet(std)
    wb.save(filename=excel_file_location)

    with ExcelWriter(excel_file_location, engine="openpyxl", mode="a") as writer:
        updated_profile.to_excel(writer, sheet_name="FDP")
        writer.save()
        writer.close()

    pickle.dump(
        updated_profile, open(result_directory + "/PROFILES/FDP_Inputs.p", "wb")
    )

    return


def create_buffer_profile(
    start: datetime = None,
    end: datetime = None,
    resolution: timedelta = None,
    prediction_delta: timedelta = None,
    seed_value: int = None,
    EMS: str = None,
    window_size_between: tuple = None,
    sample_rate: float = None,
    charging_capacity: float = None,
    result_directory: str = None,
    excel_file_location: ExcelFile = None,
    excel_file_name: str = None,
    pickled_profile: str = None,
    show_preview: bool = False,
) -> DataFrame:

    """ 
        Creates buffer profiles for storage devices (for one EMS) and 
        stores it as an excel sheet in the scenario input file.

        Parameters
        ----------
        start, end, resolution : datetime
            Simulation time related parameter

        window_size_between : int
            Used for seed generators

        EMS: str
            Contains EMS name for input file storage
        
        charging_capacity: float
            Max power capacity of the buffer per period

        window_size_between: tuple
            Used for window selection 
        
        sample_rate: float
            Used window selection (as input for df.sample(frac=sample_rate))

        input_file: str
            Contains path to excel input file

        buffer_profile_path: str
            Contains path to pickled profile
        
        Returns
        -------
        None
    """

    # Remove old data from excel
    sheet_name = EMS + ("_Storage")

    try:

        wb = openpyxl.load_workbook(excel_file_location)
        std = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(std)
        wb.save(filename=excel_file_location)

    except:
        pass

    finally:

        # No stored profile, create a new one
        if pickled_profile is None:

            assert charging_capacity > 0, "Peak power needs to be a positive number"

            # Get device constraints data structure over total simulation time
            buffer_profile = completely_unconstrained_profile(
                start=start, end=end, resolution=resolution
            )

            # Create a profile for each day of the simulation horizon
            for day in date_range(start=start, end=end, freq="D"):

                # Create an empty device constraint profile for one day
                df = completely_unconstrained_profile(
                    start=day, end=day + timedelta(days=1), resolution=resolution
                )

                # Remove overlapping indices
                df.drop(df.index[-1], inplace=True)

                # Assign buffer capacity
                max_capacity_per_timeslot = (
                    charging_capacity * resolution.seconds / 3600
                )

                # Set "max" values on a random selection of indices
                numpy_seed(seed_value)
                seed(seed_value)

                # Create random charging periods
                for index in (
                    df[: -(2 * int(prediction_delta / resolution))]
                    .sample(frac=sample_rate, random_state=seed_value)
                    .index
                ):
                    df.loc[
                        index : index
                        + timedelta(minutes=resolution.seconds / 60)
                        * randint(window_size_between[0], window_size_between[1]),
                        "max",
                    ] = round(uniform(1, max_capacity_per_timeslot))

                df.iloc[0, :] = nan
                df.iloc[-1, :] = nan

                # Filter out periods with less than 2 steps
                for index in df.groupby("max").filter(lambda x: len(x) < 2).index:
                    df.loc[index, "max"] = nan

                for index in df.loc[df["max"] > 0].index:
                    df.loc[index, "derivative max"] = max_capacity_per_timeslot
                    df.loc[index, "derivative min"] = 0

                _sum = 0
                window_start_times = list()
                window_end_times = list()
                window_storage_capacity = list()

                # Aggregate buffer window values and store window timestamps
                for index, i in zip(df["max"].index, range(df["max"].size)):

                    # Skip at first row
                    if i == 0:
                        continue

                    # Save _sum if greater than zero at last row (if it's a window)
                    if i == df["max"].size - 1:
                        if _sum > 0 and df["derivative max"].iloc[i] > 0:
                            df["max"].iloc[i] = _sum
                            window_end_times.append(index.to_pydatetime())
                        continue

                    # If row value is not nan
                    if not isnull(df["max"].iloc[i]):

                        # Add buffer window values..
                        _sum += df["max"].iloc[i]

                        # ..then overwrite with nan
                        df["max"].iloc[i] = nan

                        # If previous value is not nan
                        if isnull(df["derivative max"].iloc[i - 1]):
                            window_start_times.append(index.to_pydatetime())

                        # If next value is nan store _sum at actual index and save window end datetime
                        if isnull(df["max"].iloc[i + 1]):
                            df["max"].iloc[i] = _sum
                            window_end_times.append(index.to_pydatetime())
                            window_storage_capacity.append(_sum)

                # Store window data
                window_data_tuples = list(
                    zip(window_start_times, window_end_times, window_storage_capacity)
                )

                # Populate windows
                start_fill = 0

                cnt = 1
                for (
                    window_start,
                    window_end,
                    window_storage_capacity,
                ) in window_data_tuples:

                    buffer_storage_capacity = window_storage_capacity - start_fill

                    fill_between_lower_bound = time(
                        hour=window_start.hour,
                        minute=window_start.minute,
                        second=window_start.second,
                    )
                    fill_between_upper_bound = time(
                        hour=window_end.hour,
                        minute=window_end.minute,
                        second=window_end.second,
                    )

                    min_fill_duration = timedelta(
                        hours=buffer_storage_capacity / charging_capacity
                    )

                    if (
                        time_duration(
                            fill_between_lower_bound, fill_between_upper_bound
                        )
                        < min_fill_duration
                    ):
                        raise ValueError(
                            "Not enough time to completely fill buffer. Try again or modify sample_rate or window_size_between"
                        )

                    min_fill_steps = min_fill_duration // resolution
                    min_empty_time = fill_between_lower_bound
                    min_full_time = time_plus(
                        fill_between_lower_bound, min_fill_duration
                    )
                    max_empty_time = time_plus(
                        fill_between_upper_bound, -min_fill_duration
                    )
                    max_full_time = fill_between_upper_bound

                    for d in date_range(window_start, window_end):
                        end_fill = start_fill + buffer_storage_capacity
                        min_empty_datetime = datetime.combine(d, min_empty_time)
                        min_full_datetime = datetime.combine(d, min_full_time)
                        max_empty_datetime = datetime.combine(d, max_empty_time)
                        max_full_datetime = datetime.combine(d, max_full_time)
                        df["max"].loc[
                            window_start : min_empty_datetime - resolution
                        ] = start_fill
                        df["min"].loc[
                            window_start : max_empty_datetime - resolution
                        ] = start_fill

                        for fill_step in range(min_fill_steps):

                            # Fill max column forward from min_fill_datetime
                            fill_max_value = start_fill + (
                                fill_step + 1
                            ) * charging_capacity * resolution / timedelta(hours=1)
                            df["max"].loc[
                                min_empty_datetime
                                - resolution
                                + (fill_step + 1) * resolution
                            ] = fill_max_value

                            # Fill min column backward from max_fill_datetime
                            fill_min_value = end_fill - (
                                fill_step + 1
                            ) * charging_capacity * resolution / timedelta(hours=1)
                            df["min"].loc[
                                max_full_datetime
                                - resolution
                                - (fill_step + 1) * resolution
                            ] = fill_min_value

                        df["max"].loc[
                            min_full_datetime - resolution : window_end
                        ] = end_fill
                        df["min"].loc[
                            max_full_datetime - resolution : window_end
                        ] = end_fill

                        # Update for next window
                        start_fill = end_fill

                for index, i in zip(df["max"].index, range(df["max"].size)):
                    if not isnull(df["max"].iloc[i]) and isnull(df["max"].iloc[i + 1]):
                        df.iloc[i, :] = nan

                df["derivative max"].fillna(0, inplace=True)
                df["derivative min"].fillna(0, inplace=True)

                buffer_profile.loc[df.index[0] : df.index[12], :] = df.shift(-1)


        else:

            buffer_profile = pickle.load(open(pickled_profile, "rb"))

            assert (
                buffer_profile.index[-1] >= end
            ), "Pickled profile does not match simulation runtime"

    with ExcelWriter(excel_file_location, engine="openpyxl", mode="a") as writer:
        buffer_profile.to_excel(writer, sheet_name=sheet_name)
        writer.save()
        writer.close()

    pickle.dump(
        buffer_profile, open(result_directory + "/PROFILES/" + sheet_name + ".p", "wb")
    )

    if show_preview:
        fig = plt.figure(2, figsize=(20, 10))
        fig.suptitle("{}: Buffer profile".format(EMS), fontsize=14)

        ax1 = plt.subplot(2, 1, 1)

        buffer_profile["min"].plot(
            kind="line", title="Buffer SOC constraints", grid=True, ax=ax1, c="black"
        )
        buffer_profile["max"].plot(
            kind="line", title="Buffer constraints", grid=True, ax=ax1, c="black"
        )
        plt.setp(ax1.get_xticklabels(), visible=False)

        ax2 = plt.subplot(2, 1, 2)
        buffer_profile["derivative max"].plot(
            kind="line", title="Buffer charging constraint", grid=True, ax=ax2
        )
        plt.setp(ax2.get_xticklabels(), visible=False)

        plt.show()

    # Basic stats
    total_days = (end - start).days
    day_cnt = 0

    return


def create_solar_generator_profile(
    start: datetime = None,
    end: datetime = None,
    resolution: timedelta = None,
    EMS: str = None,
    seed_value: int = None,
    histogram_parameters: tuple = None,
    peak_power: float = None,
    dispatch_factor: float = None,
    result_directory: str = None,
    excel_file_location: ExcelFile = None,
    excel_file_name: str = None,
    copy_from_excel: bool = False,
    pickled_profile: str = None,
    show_preview: bool = False,
    repeat_daily: bool = False,
) -> None:

    """ 
        Creates solar profiles for PV generators from histograms (for one EMS) and 
        stores it as an excel sheet in the scenario input file.

        Parameters
        ----------
        start, end, resolution : datetime
            Simulation time related parameter

        seed_value : int
            Used for seed generators

        EMS: str
            Contains EMS name for input file storage

        peak_power: float (positive sign)
            Max power capacity of the solar generator per period

        histogram_parameters: tuple
            Mean and std for histogram function
                        
        dispatch_factor : float
            Defines wheter a PV power can be curtailed or not
            If 1 is passed, PV generator is fully curtailable

        input_file: str
            Contains path to excel input file

        generator_profile_path: str
            Contains path to pickled profile

        Returns
        -------
        None
    """
    # Remove old data from excel
    sheet_name = EMS + ("_Generator")

    try:
        wb = openpyxl.load_workbook(excel_file_location)
        std = wb.get_sheet_by_name(sheet_name)

        if not copy_from_excel:
            wb.remove_sheet(std)
            wb.save(filename=excel_file_location)
        else:
            sheet_available = True
    except:
        sheet_available = False

    finally:

        # If no path to a pickled profile has been passed, generate a new profile
        if pickled_profile is None:

            # Assertations
            assert (
                dispatch_factor >= 0 and dispatch_factor <= 1
            ), "Dispatch factor needs to be a number between 0 an 1"
            assert peak_power >= 0, "Peak power needs to be a positive number"

            # Flag
            repeat = repeat_daily

            # Set simulation index from attributes
            simulation_time = date_range(start=start, end=end, freq=resolution)

            # Create full simulation time dataframe
            fdp_full_time_profile = completely_unconstrained_profile(
                start=start, end=end, resolution=resolution
            )

            fdp_full_time_profile.index.name = "datetime"

            if copy_from_excel and sheet_available:

                # Copy from existing sheet
                solar_profile = read_excel(
                    io=excel_file_location,
                    sheet_name=EMS + "_Generator",
                    header=0,
                    squeeze=True, 
                )

                # # Remove existing sheet
                wb = openpyxl.load_workbook(excel_file_location)
                std = wb.get_sheet_by_name(sheet_name)
                wb.remove_sheet(std)
                wb.save(filename=excel_file_location)

                solar_profile.set_index("datetime", inplace=True)

                # Infer and assign index resolution
                solar_profile.index.freq = infer_freq(solar_profile.index)

                assert (
                    to_timedelta(solar_profile.index.freq) == resolution
                ), "Resolution not equal"

                # Set repeat if length is not equal
                if len(solar_profile) < len(simulation_time):
                    repeat = True

                if repeat:

                    delta_profile = solar_profile.index[-1] - solar_profile.index[0]

                    if delta_profile > timedelta(days=1):
                        solar_profile = solar_profile.loc[
                            start : start + timedelta(days=1)
                        ]

                    simulation_days = (end - start).days

                    if not delta_profile < timedelta(days=1):

                        delta_profile = timedelta(days=1)

                    for i in range(0, simulation_days):
                        day_start = start + timedelta(days=i)
                        fdp_full_time_profile.loc[
                            day_start : day_start + delta_profile
                        ] = solar_profile.loc[start : start + delta_profile].values

                    solar_profile = fdp_full_time_profile

            else:

                # Set full year index for python enlopy value generator
                full_year_index = date_range(
                    start=datetime(year=2018, month=1, day=1),
                    end=datetime(year=2019, month=1, day=1, hour=0),
                    freq=resolution,
                )
                # Drop last index (00:00)
                full_year_index = full_year_index.drop(full_year_index[-1])

                # Fix random parameters
                numpy_seed(randint(0, 1000))
                seed(randint(0, 1000))

                if resolution == timedelta(minutes=15):

                    # Mean and std for histogram function
                    mu, sigma = histogram_parameters
                    histogram_input_values = normal(mu, sigma, 5000)
                    count, bins, ignored = hist(
                        histogram_input_values, 56, density=False
                    )
                    _ = [b.remove() for b in ignored]

                    # Create series for daily profile and store histogram function values
                    daily_profile = Series(
                        index=RangeIndex(start=1, stop=97, step=1), data=0
                    )
                    daily_profile[24:80] = count

                    # Python enlopy conversion function
                    _daily_profile = enlopy.utils.make_timeseries(
                        x=daily_profile,
                        year=2018,
                        length=96,
                        startdate=start,
                        freq=resolution,
                    )

                    # Add seasonality
                    yearly_curve = (
                        -cos(2 * pi / 35040 * linspace(0, 35039, 35040) + 0.2) * 50
                        + 100
                    ) * 0.75

                    yearly_profiles = []

                    # Create profiles for each day with randomized starting offsets
                    for i in range(1, 366):
                        # numpy_seed(randint(0,1000))
                        # seed(randint(0,1000))
                        offset = randint(-4, 6)
                        daily_profile = _daily_profile.copy()
                        daily_profile[24 + offset : 80 + offset] = add_noise(
                            daily_profile[24:80] * (random() + 0.5) / 2,
                            mode=2,
                            st=0.5,
                            Lmin=0,
                            r=0,
                            freq=resolution,
                        ).values
                        yearly_profiles.append(list(daily_profile))

                elif resolution == timedelta(minutes=60):

                    # Mean and std for histogram function
                    mu, sigma = histogram_parameters
                    histogram_input_values = normal(mu, sigma, 5000)
                    count, bins, ignored = hist(
                        histogram_input_values, 11, density=False
                    )
                    _ = [b.remove() for b in ignored]

                    # Create series for daily profile and store histogram function values
                    daily_profile = Series(
                        index=RangeIndex(start=1, stop=25, step=1), data=0
                    )
                    daily_profile[7:18] = count

                    # Python enlopy conversion function
                    _daily_profile = enlopy.utils.make_timeseries(
                        x=daily_profile,
                        year=2018,
                        length=24,
                        startdate=start,
                        freq=resolution,
                    )

                    # Add seasonality
                    yearly_curve = (
                        -cos(2 * pi / 8760 * linspace(0, 8759, 8760) + 0.2) * 50 + 100
                    ) * 0.75

                    yearly_profiles = []

                    # Create profiles for each day with randomized starting offsets
                    for i in range(1, 366):
       
                        offset = randint(-1, 1)
                        daily_profile = _daily_profile.copy()
                        daily_profile[7 + offset : 18 + offset] = add_noise(
                            daily_profile[7:18] * (random() + 0.5) / 2,
                            mode=2,
                            st=0.5,
                            Lmin=0,
                            r=0,
                            freq=resolution,
                        ).values
                        yearly_profiles.append(list(daily_profile))

                yearly_profiles = [
                    item for sublist in yearly_profiles for item in sublist
                ]
                full_year_profile = yearly_profiles * yearly_curve
                full_year_profile = full_year_profile / full_year_profile.max()
                generator_profile = Series(
                    data=full_year_profile, index=full_year_index
                )
                generator_profile *= peak_power

                # Get device constraints datastructure
                solar_profile = limited_production_profile(
                    start=start, end=end, resolution=resolution, capacity=peak_power
                )

                # Store profile values
                solar_profile["derivative equals"] = around(
                    -generator_profile.loc[start:end].values, 1
                )

                # Convert if curtailabilty is specified
                if dispatch_factor is not None:
                    if dispatch_factor == 0:
                        pass

                    if dispatch_factor == 1:
                        solar_profile["derivative min"] = solar_profile[
                            "derivative equals"
                        ]
                        solar_profile["derivative equals"] = nan

                    else:
                        solar_profile["derivative min"] = solar_profile[
                            "derivative equals"
                        ]
                        solar_profile["derivative equals"] = solar_profile[
                            "derivative equals"
                        ] * (1 - dispatch_factor)

                if repeat:

                    delta_profile = (
                        fdp_full_time_profile.index[-1] - fdp_full_time_profile.index[0]
                    )

                    simulation_days = (end - start).days

                    if not delta_profile < timedelta(days=1):

                        delta_profile = timedelta(days=1)

                    for i in range(0, simulation_days):
                        day_start = start + timedelta(days=i)
                        fdp_full_time_profile.loc[
                            day_start : day_start + delta_profile
                        ] = solar_profile.loc[start : start + delta_profile].values

                    solar_profile = fdp_full_time_profile

        # Use a pickled profile in case a profile path has been specified
        else:

            solar_profile = pickle.load(open(pickled_profile, "rb"))

            assert (
                solar_profile.index[-1] >= end
            ), "Pickled profile does not match simulation runtime"

        if not isnan(solar_profile["derivative equals"].min()):
            limit = solar_profile["derivative equals"].min()
            profile = solar_profile["derivative equals"]
        else:
            limit = solar_profile["derivative min"].min()
            profile = solar_profile["derivative min"]

        # Write to input file
        with ExcelWriter(excel_file_location, engine="openpyxl", mode="a") as writer:
            solar_profile.to_excel(writer, sheet_name=EMS + ("_Generator"))
            writer.save()
            writer.close()

        pickle.dump(
            solar_profile,
            open(result_directory + "/PROFILES/" + sheet_name + ".p", "wb"),
        )

        # Plot profiles
        if show_preview:

            fig = plt.figure(1, figsize=(20, 10))
            fig.suptitle("{} PV profile".format(EMS), fontsize=14)
            ax = plt.subplot(1, 1, 1)
            ax.set_ylim(limit, 0)
            profile.plot(kind="Line", grid=True, ax=ax)

            plt.show()

        # Basic stats
        total_days = (end - start).days
        day_cnt = 0
        for day in date_range(start=start, end=end, freq="D"):

            # Exit condition
            if day_cnt > total_days - 1:
                break

            # Show the daily energy production
            pv_energy = around(
                profile.loc[start : start + timedelta(days=1)].sum()
                * (resolution / timedelta(minutes=60)),
                2,
            )
            print("Day {} | Date: {} | PV energy: {}".format(day_cnt, day, pv_energy))

            day_cnt += 1
        return


def create_grid_constraints(
    result_directory: str,
    excel_file_location: str,
    excel_file_name: str,
    EMS: str,
    start: datetime,
    end: datetime,
    resolution: datetime,
    power_capacity_max: int,
    feedin_capacity_max: int,
    pickled_profile: str = None,
    show_preview: bool = False,
) -> None:
    """ 
        Creates solar profiles for PV generators from histograms (for one EMS) and 
        stores it as an excel sheet in the scenario input file.

        Parameters
        ----------
        start, end, resolution : datetime
            Simulation time related parameter

        seed_value : int
            Used for seed generators

        EMS: str
            Contains EMS name for input file storage

        power_capacity_max, feedin_capacity_max: float (positive signs)
            Max power capacity from and to grid for EMS in total

        input_file: str
            Contains path to excel input file

        Returns
        -------
        None
    """
    # Assertations
    assert (
        power_capacity_max >= 0 and feedin_capacity_max >= 0
    ), "Power values needs to be positive numbers"

    df = completely_unconstrained_profile(start=start, end=end, resolution=resolution)

    df["derivative max"] = power_capacity_max
    df["derivative min"] = -feedin_capacity_max

    # Remove old data from excel
    sheet_name = EMS + ("_Grid")
    try:
        wb = openpyxl.load_workbook(excel_file_location)
        std = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(std)
        wb.save(filename=excel_file_location)

    except:
        pass

    finally:

        if pickled_profile is not None:

            df = pickle.load(open(pickled_profile, "rb"))
            assert (
                df.index[-1] >= end
            ), "Pickled profile does not match simulation runtime"

        else:
            pass

        with ExcelWriter(excel_file_location, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name=sheet_name)
            writer.save()
            writer.close()

        pickle.dump(df, open(result_directory + "/PROFILES/" + sheet_name + ".p", "wb"))

    # Basic stats
    total_days = (end - start).days
    day_cnt = 0


    return
